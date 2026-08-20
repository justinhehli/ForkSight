import io
import json
import os
import subprocess
import sys
import threading
import time
from enum import Enum
from pathlib import Path

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font
from PIL import Image


def _find_repo_root() -> Path:
    start = Path(__file__).resolve()
    for parent in [start] + list(start.parents):
        if (parent / ".git").exists():
            return parent
    raise RuntimeError("No repo root ('.git' folder) found")


REPO_ROOT = _find_repo_root()
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from fastapi import BackgroundTasks, FastAPI, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import Response
from pydantic import BaseModel

from AnnotationTool.backend.bwrap_util import RESTART_EXIT_CODE, is_sandboxed
from AnnotationTool.backend.pipeline.discovery import (
    IS_TRAIN_ENV,
    PROJECTS_PARENT_DIR,
    SEGMENTATION_DIR_NAME,
    TOOL_ENVIRONMENT,
    TRAIN_PROJECT_NAME,
    find_project_tiles,
    fork_detection_dir,
    list_candidate_dirs,
    load_pipeline_settings,
    load_registered_projects,
    resolve_unc_path,
    save_pipeline_settings,
    save_registered_projects,
)
from AnnotationTool.backend.pipeline.annotations_store import (
    PipelineStatus,
    load_annotations,
    pipeline_log_path,
    save_annotations,
)
from AnnotationTool.backend.pipeline.process_util import is_pid_running, terminate_process_tree
from AnnotationTool.backend.pipeline.progress_util import clear_progress, read_progress
from AnnotationTool.backend.pipeline.run_pipeline import cleanup_stale_temp_dirs
from AnnotationTool.backend.util import get_repo_root
from Segmentation.PreProcessing.General.tif_to_png import convert_tif_to_png

app = FastAPI(title="ForkSight Annotator API")

app.add_middleware(
    CORSMiddleware,
    # allow_origins=["http://localhost:5173"],
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# models
# ====================


class JunctionType(str, Enum):
    ReplicationFork50 = "Replication Fork 50%"
    ReplicationFork100 = "Replication Fork 100%"
    ReversedFork50 = "Reversed Fork 50%"
    ReversedFork100 = "Reversed Fork 100%"


class PipelineMode(str, Enum):
    # Tiles are randomly sampled and processed one at a time (segment, save,
    # detect, persist) until enough total junctions have been found.
    Sequential = "sequential"
    # Two-stage pipeline: segment a batch of tiles, then detect junctions in
    # all of them, only persisting once the whole batch is done.
    Staged = "staged"


# Weight that each fork label contributes towards the fork-ratio calculation.
FORK_WEIGHTS = {
    JunctionType.ReplicationFork50: 0.5,
    JunctionType.ReplicationFork100: 1.0,
    JunctionType.ReversedFork50: 0.5,
    JunctionType.ReversedFork100: 1.0,
}

REPLICATION_FORK_LABELS = {
    JunctionType.ReplicationFork50, JunctionType.ReplicationFork100}
REVERSED_FORK_LABELS = {
    JunctionType.ReversedFork50, JunctionType.ReversedFork100}

BUILTIN_LABELS = {member.value for member in JunctionType}


class Point(BaseModel):
    id: str
    x: int
    y: int
    labels: list[str]


class ImageAnnotations(BaseModel):
    processed: bool
    points: list[Point]


class ProjectSelection(BaseModel):
    names: list[str]


class CreateLabelWrapper(BaseModel):
    label: str


class PipelineSettingsModel(BaseModel):
    pipeline_mode: str
    sequential_target_junction_count: int
    staged_sample_count: int


# concurrency
# ====================
# One pipeline may run at a time globally, as determined from annotations.json
# This lock only makes the check-then-set sequence atomic within this backend process
_pipeline_start_lock = threading.Lock()

# Per-project locks guard read-modify-write of annotations.json.
_project_locks: dict[str, threading.Lock] = {}
_project_locks_guard = threading.Lock()


def _get_project_lock(project: str) -> threading.Lock:
    with _project_locks_guard:
        if project not in _project_locks:
            _project_locks[project] = threading.Lock()
        return _project_locks[project]


# helpers
# ====================

def project_dir(project: str) -> Path:
    if IS_TRAIN_ENV:
        if project != TRAIN_PROJECT_NAME or not PROJECTS_PARENT_DIR.is_dir():
            raise HTTPException(404, f"Project '{project}' not found")
        return PROJECTS_PARENT_DIR

    p = PROJECTS_PARENT_DIR / project
    registered = load_registered_projects(PROJECTS_PARENT_DIR)
    if project not in registered or not p.is_dir():
        raise HTTPException(404, f"Project '{project}' not found")
    return p


# routes
# ====================


@app.get("/environment")
def get_environment() -> dict:
    return {"environment": TOOL_ENVIRONMENT.value}


@app.get("/projects")
def list_projects() -> list[str]:
    if not PROJECTS_PARENT_DIR.exists():
        raise HTTPException(
            500, f"PROJECTS_PARENT_DIR does not exist: {PROJECTS_PARENT_DIR}")
    if IS_TRAIN_ENV:
        return [TRAIN_PROJECT_NAME]
    registered = load_registered_projects(PROJECTS_PARENT_DIR)
    return sorted(name for name in registered if (PROJECTS_PARENT_DIR / name).is_dir())


@app.get("/project-candidates")
def get_project_candidates() -> list[dict]:
    if IS_TRAIN_ENV:
        raise HTTPException(
            400, "Project discovery is not applicable in the TRAIN environment")
    if not PROJECTS_PARENT_DIR.exists():
        raise HTTPException(
            500, f"PROJECTS_PARENT_DIR does not exist: {PROJECTS_PARENT_DIR}")
    return list_candidate_dirs(PROJECTS_PARENT_DIR)


def _exit_for_restart(delay: float = 0.5) -> None:
    time.sleep(delay)
    print(f"[main] Exiting with code {RESTART_EXIT_CODE} so the supervisor "
          "relaunches with a fresh sandbox", file=sys.stderr, flush=True)
    os._exit(RESTART_EXIT_CODE)


@app.post("/project-candidates")
def set_project_candidates(selection: ProjectSelection, background_tasks: BackgroundTasks) -> list[dict]:
    if IS_TRAIN_ENV:
        raise HTTPException(
            400, "Project registration is not applicable in the TRAIN environment")

    previously_registered = load_registered_projects(PROJECTS_PARENT_DIR)
    save_registered_projects(PROJECTS_PARENT_DIR, selection.names)
    newly_registered = set(selection.names) - previously_registered

    print(f"[main] set_project_candidates: sandboxed={is_sandboxed()} "
          f"newly_registered={newly_registered}", file=sys.stderr, flush=True)

    # A project's AutomaticForkDetection dir is only created once it's actually
    # registered here, not merely discovered as a candidate.
    if is_sandboxed():
        # This process can't create the dir itself: the project's folder isn't
        # in the sandbox's read-write allowlist until the backend is relaunched
        # with a fresh one, which is exactly what needs to happen here.
        missing_dirs = [name for name in newly_registered
                        if not fork_detection_dir(PROJECTS_PARENT_DIR / name).is_dir()]
        print(f"[main] missing_dirs={missing_dirs}",
              file=sys.stderr, flush=True)
        if missing_dirs:
            background_tasks.add_task(_exit_for_restart)
    else:
        for name in newly_registered:
            fork_detection_dir(PROJECTS_PARENT_DIR / name).mkdir(
                parents=True, exist_ok=True)

    return list_candidate_dirs(PROJECTS_PARENT_DIR)


@app.get("/pipeline-settings")
def get_pipeline_settings() -> dict:
    return load_pipeline_settings(PROJECTS_PARENT_DIR)


@app.post("/pipeline-settings")
def set_pipeline_settings(settings: PipelineSettingsModel) -> dict:
    if settings.pipeline_mode not in {m.value for m in PipelineMode}:
        raise HTTPException(400, f"Invalid mode '{settings.pipeline_mode}'")
    if settings.sequential_target_junction_count <= 0:
        raise HTTPException(
            400, "sequential_target_junction_count must be positive")
    if settings.staged_sample_count <= 0:
        raise HTTPException(
            400, "staged_sample_count must be positive")

    save_pipeline_settings(
        PROJECTS_PARENT_DIR, settings.pipeline_mode, settings.sequential_target_junction_count,
        settings.staged_sample_count)
    return load_pipeline_settings(PROJECTS_PARENT_DIR)


@app.get("/project-candidates/{name:path}/path")
def get_project_candidate_path(name: str):
    if IS_TRAIN_ENV:
        if name != TRAIN_PROJECT_NAME or not PROJECTS_PARENT_DIR.is_dir():
            raise HTTPException(404, "Directory not found")
        return {"path": str(resolve_unc_path(PROJECTS_PARENT_DIR.resolve()))}

    target = (PROJECTS_PARENT_DIR / name).resolve()
    try:
        target.relative_to(PROJECTS_PARENT_DIR.resolve())
    except ValueError:
        raise HTTPException(400, "Invalid project path")
    if not target.is_dir():
        raise HTTPException(404, "Directory not found")

    return {"path": str(resolve_unc_path(target))}


@app.get("/projects/{project:path}/images")
def list_images(project: str):
    pd = project_dir(project)
    ann = load_annotations(pd)
    images = [
        {"id": image_id, "name": img["display_name"],
            "processed": img.get("processed", False),
            "archived": img.get("archived", False)}
        for image_id, img in ann["images"].items()
    ]
    images.sort(key=lambda i: i["name"])
    return images


@app.get("/projects/{project:path}/images/{image_id}")
def serve_image(project: str, image_id: str):
    pd = project_dir(project)
    ann = load_annotations(pd)
    img_ann = ann["images"].get(image_id)
    if img_ann is None:
        raise HTTPException(404, "Image not found")

    tif_path = pd / img_ann["source_tif"]
    if not tif_path.exists():
        raise HTTPException(404, "Source TIF not found")

    png = convert_tif_to_png(tif_path)
    buf = io.BytesIO()
    png.save(buf, format="PNG")
    return Response(content=buf.getvalue(), media_type="image/png",
                    headers={"Cache-Control": "public, max-age=31536000, immutable"})


@app.get("/projects/{project:path}/images/{image_id}/mask")
def serve_mask(project: str, image_id: str, request: Request):
    MASK_OVERLAY_COLOR = (0, 255, 255, 130)

    pd = project_dir(project)
    mask_path = fork_detection_dir(
        pd) / SEGMENTATION_DIR_NAME / f"{image_id}.png"
    if not mask_path.is_file():
        raise HTTPException(404, "Segmentation mask not found")

    # Unlike the source image, mask can be regenerated by re-running fork detection
    # -> key the cache off its mtime so a stale cached copy is never served
    etag = f'"{image_id}-{int(mask_path.stat().st_mtime)}"'
    headers = {"ETag": etag,
               "Cache-Control": "private, max-age=0, must-revalidate"}
    if request.headers.get("if-none-match") == etag:
        return Response(status_code=304, headers=headers)

    mask = np.array(Image.open(mask_path).convert("L"))
    overlay = np.zeros((*mask.shape, 4), dtype=np.uint8)
    overlay[mask > 127] = MASK_OVERLAY_COLOR

    buf = io.BytesIO()
    Image.fromarray(overlay, mode="RGBA").save(buf, format="PNG")
    return Response(content=buf.getvalue(), media_type="image/png", headers=headers)


@app.get("/projects/{project:path}/annotations")
def get_annotations(project: str) -> dict:
    return load_annotations(project_dir(project))


@app.put("/projects/{project:path}/annotations/{image_id}", status_code=204)
def save_image_annotations(project: str, image_id: str, data: ImageAnnotations):
    pd = project_dir(project)
    with _get_project_lock(project):
        ann = load_annotations(pd)
        if image_id not in ann["images"]:
            raise HTTPException(404, f"Image '{image_id}' not found")
        ann["images"][image_id].update(data.model_dump())
        save_annotations(pd, ann)


@app.post("/projects/{project:path}/images/{image_id}/archive", status_code=204)
def archive_image(project: str, image_id: str):
    pd = project_dir(project)
    with _get_project_lock(project):
        ann = load_annotations(pd)
        if image_id not in ann["images"]:
            raise HTTPException(404, f"Image '{image_id}' not found")
        ann["images"][image_id]["archived"] = True
        save_annotations(pd, ann)


@app.post("/projects/{project:path}/images/{image_id}/unarchive", status_code=204)
def unarchive_image(project: str, image_id: str):
    pd = project_dir(project)
    with _get_project_lock(project):
        ann = load_annotations(pd)
        if image_id not in ann["images"]:
            raise HTTPException(404, f"Image '{image_id}' not found")
        ann["images"][image_id]["archived"] = False
        save_annotations(pd, ann)


@app.post("/projects/{project:path}/labels")
def add_custom_label(project: str, body: CreateLabelWrapper) -> list[str]:
    label = body.label.strip()
    if not label:
        raise HTTPException(400, "Label cannot be empty")
    if label in BUILTIN_LABELS:
        raise HTTPException(400, f"'{label}' is a built-in label")

    pd = project_dir(project)
    with _get_project_lock(project):
        ann = load_annotations(pd)
        additional = ann.setdefault("additional_labels", [])
        if label not in additional:
            additional.append(label)
            save_annotations(pd, ann)
        return additional


@app.delete("/projects/{project:path}/labels/{label}")
def delete_custom_label(project: str, label: str) -> list[str]:
    if label in BUILTIN_LABELS:
        raise HTTPException(
            400, f"'{label}' is a built-in label and cannot be deleted")

    pd = project_dir(project)
    with _get_project_lock(project):
        ann = load_annotations(pd)
        additional = ann.setdefault("additional_labels", [])
        if label in additional:
            additional.remove(label)
        for img_ann in ann.get("images", {}).values():
            for point in img_ann.get("points", []):
                if label in point.get("labels", []):
                    point["labels"] = [
                        l for l in point["labels"] if l != label]
        save_annotations(pd, ann)
        return additional


def _active_images(images: dict) -> dict:
    return {image_id: img for image_id, img in images.items() if not img.get("archived", False)}


def _compute_summary(images: dict) -> dict:
    # Fork counts only consider processed images - unprocessed images may still
    # be mid-annotation and would skew the counts.
    processed_points = [p for img_ann in images.values()
                        if img_ann.get("processed", False)
                        for p in img_ann.get("points", [])]
    replication_forks = sum(
        FORK_WEIGHTS.get(l, 0.0)
        for p in processed_points for l in p.get("labels", [])
        if l in REPLICATION_FORK_LABELS)
    reversed_forks = sum(
        FORK_WEIGHTS.get(l, 0.0)
        for p in processed_points for l in p.get("labels", [])
        if l in REVERSED_FORK_LABELS)
    processed_count = sum(1 for img_ann in images.values()
                          if img_ann.get("processed", False))

    return {
        "total_images": len(images),
        "processed_images": processed_count,
        "replication_fork_weighted_count": replication_forks,
        "reversed_fork_weighted_count": reversed_forks,
        "replication_reversed_ratio": round(replication_forks / reversed_forks, 3) if reversed_forks > 0 else None,
    }


# Fork labels (if set) always come first, followed by any additional labels alphabetically -
# mirrors sortLabelsForDisplay() in the frontend.
_FORK_LABEL_SET = REPLICATION_FORK_LABELS | REVERSED_FORK_LABELS


def _sort_labels_for_display(labels: list[str]) -> list[str]:
    fork_labels = [l for l in labels if l in _FORK_LABEL_SET]
    other_labels = sorted(l for l in labels if l not in _FORK_LABEL_SET)
    return fork_labels + other_labels


@app.post("/projects/{project:path}/export")
def export_project(project: str):
    ann = load_annotations(project_dir(project))
    images = ann["images"]

    export_data = {
        "summary": _compute_summary(_active_images(images)),
        "images": images,
    }

    content = json.dumps(export_data, indent=2, ensure_ascii=False)
    safe_name = project.replace("/", "_")
    return Response(
        content=content,
        media_type="application/json",
        headers={
            "Content-Disposition": f'attachment; filename="{safe_name}_annotations.json"'},
    )


@app.post("/projects/{project:path}/export-excel")
def export_project_excel(project: str):
    ann = load_annotations(project_dir(project))
    images = ann["images"]
    active_images = _active_images(images)
    summary = _compute_summary(active_images)

    wb = Workbook()
    ws = wb.active
    ws.title = "Annotations"
    bold = Font(bold=True)

    ws.append(["Summary"])
    ws.cell(row=ws.max_row, column=1).font = bold
    ws.append(["Total images", summary["total_images"]])
    ws.append(["Processed images", summary["processed_images"]])
    ws.append(["Replication fork weighted count (processed images only)",
              summary["replication_fork_weighted_count"]])
    ws.append(["Reversed fork weighted count (processed images only)",
              summary["reversed_fork_weighted_count"]])
    ws.append(["Replication / reversed ratio (processed images only)",
              summary["replication_reversed_ratio"]])
    ws.append([])

    ws.append(["Source TIF", "Processed", "X", "Y", "Labels"])
    for cell in ws[ws.max_row]:
        cell.font = bold

    for img_ann in sorted(active_images.values(), key=lambda i: i.get("display_name", "")):
        source_tif = img_ann.get("display_name", "")
        processed = "Yes" if img_ann.get("processed", False) else "No"
        points = img_ann.get("points", [])
        if not points:
            ws.append([source_tif, processed, None, None, None])
            continue
        for p in points:
            labels = ", ".join(_sort_labels_for_display(p.get("labels", [])))
            ws.append([source_tif, processed, p["x"], p["y"], labels])

    for col, width in zip("ABCDE", (45, 12, 8, 8, 40)):
        ws.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    safe_name = project.replace("/", "_")
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{safe_name}_annotations.xlsx"'},
    )


@app.get("/projects/{project:path}/pipeline-log")
def get_pipeline_log(project: str):
    pd = project_dir(project)
    log_path = pipeline_log_path(pd)
    if not log_path.is_file():
        return Response(content="", media_type="text/plain")
    return Response(content=log_path.read_text(encoding="utf-8", errors="replace"),
                    media_type="text/plain")


@app.get("/projects/{project:path}/pipeline-progress")
def get_pipeline_progress(project: str) -> dict:
    return read_progress(project_dir(project))


def _find_running_pipeline() -> str | None:
    """Return the name of a project with an active pipeline, if any.
    Projects stuck at "Running" whose recorded pipeline_runner.py process is 
    no longer alive is set to "Failed" here so it doesn't block new runs.
    """
    for name in sorted(load_registered_projects(PROJECTS_PARENT_DIR)):
        pd = PROJECTS_PARENT_DIR / name
        if not pd.is_dir():
            continue
        ann = load_annotations(pd)
        if ann.get("junction_detection_pipeline_status") != PipelineStatus.Running:
            continue
        if is_pid_running(ann.get("pipeline_pid")):
            return name
        with _get_project_lock(name):
            ann = load_annotations(pd)
            if ann.get("junction_detection_pipeline_status") == PipelineStatus.Running:
                ann["junction_detection_pipeline_status"] = PipelineStatus.Failed
                ann["pipeline_error"] = (
                    "Pipeline process is no longer running "
                    "(backend restarted or the process crashed)."
                )
                ann["pipeline_pid"] = None
                save_annotations(pd, ann)
                clear_progress(pd)
    return None


@app.get("/pipeline-status")
def get_pipeline_status():
    return {"running_project": _find_running_pipeline()}


@app.post("/projects/{project:path}/stop-junction-detection")
def stop_junction_detection(project: str):
    pd = project_dir(project)

    with _get_project_lock(project):
        ann = load_annotations(pd)
        if ann.get("junction_detection_pipeline_status") != PipelineStatus.Running:
            raise HTTPException(
                409, f"No junction detection pipeline is running for project '{project}'")
        pid = ann.get("pipeline_pid")

    # Terminating the process tree can take a few seconds (graceful attempt
    # before escalating to a forceful kill) - do this outside the lock.
    terminate_process_tree(pid)
    cleanup_stale_temp_dirs()

    with _get_project_lock(project):
        ann = load_annotations(pd)
        ann["junction_detection_pipeline_status"] = PipelineStatus.Failed
        ann["pipeline_error"] = "Pipeline stopped by user."
        ann["pipeline_pid"] = None
        save_annotations(pd, ann)
        clear_progress(pd)

    return {"status": PipelineStatus.Failed, "project": project}


def _count_total_junctions(images: dict) -> float:
    return sum(
        FORK_WEIGHTS.get(l, 0.0)
        for img in images.values()
        for p in img.get("points", [])
        for l in p.get("labels", [])
        if l in REPLICATION_FORK_LABELS or l in REVERSED_FORK_LABELS
    )


class RunDetectionRequest(BaseModel):
    # Meaning depends on the project's pipeline mode:
    #   sequential -> additional total junctions to find this run
    #   staged     -> number of the project's tiles to sample in this run
    amount: float | None = None


@app.post("/projects/{project:path}/run-junction-detection")
def run_junction_detection(project: str, body: RunDetectionRequest = RunDetectionRequest()):
    pd = project_dir(project)

    with _pipeline_start_lock:
        running_project = _find_running_pipeline()
        if running_project is not None:
            raise HTTPException(
                409, f"A junction detection pipeline is already running for project '{running_project}'")

        global_settings = load_pipeline_settings(PROJECTS_PARENT_DIR)

        with _get_project_lock(project):
            ann = load_annotations(pd)

            if IS_TRAIN_ENV:
                # TRAIN environment always runs staged over the full set of tiles
                mode = PipelineMode.Staged
                mode_args = ["--sample-count",
                             str(len(find_project_tiles(pd)))]
            else:
                # mode of project (if set before) trumps global default setting
                mode = ann.get(
                    "pipeline_mode") or global_settings["pipeline_mode"]

                if mode == PipelineMode.Sequential:
                    additional = body.amount if body.amount is not None else global_settings[
                        "sequential_target_junction_count"]
                    if additional is None or additional <= 0:
                        raise HTTPException(
                            400, "amount (additional junctions) must be positive")
                    target_junction_count = int(round(
                        _count_total_junctions(_active_images(ann["images"])) + additional))
                    mode_args = ["--target-junction-count",
                                 str(target_junction_count)]
                else:
                    count = body.amount if body.amount is not None else global_settings[
                        "staged_sample_count"]
                    if count is None or count <= 0:
                        raise HTTPException(
                            400, "amount (number of tiles) must be positive")
                    mode_args = ["--sample-count", str(int(round(count)))]

            ann["pipeline_mode"] = mode
            ann["junction_detection_pipeline_status"] = PipelineStatus.Running
            ann["pipeline_error"] = None
            save_annotations(pd, ann)

        # Spawn the pipeline as its own OS process, this request returns immediately
        # and pipeline_runner.py itself updates annotations.json when it's done
        log_path = pipeline_log_path(pd)
        log_path.parent.mkdir(parents=True, exist_ok=True)
        clear_progress(pd)
        log_file = open(log_path, "w", encoding="utf-8")
        env = os.environ.copy()
        env["PYTHONUNBUFFERED"] = "1"

        # this subprocess inherits the backend's sandbox (bwrap) restrictions
        cmd = [
            sys.executable, "-u", "-m", "AnnotationTool.backend.pipeline.pipeline_runner",
            "--project-dir", str(pd),
            "--mode", mode,
            *mode_args,
        ]
        try:
            proc = subprocess.Popen(
                cmd,
                cwd=str(get_repo_root()),
                env=env,
                stdout=log_file,
                stderr=subprocess.STDOUT,
                # Puts pipeline_runner.py and its own worker subprocesses in one
                # process group, so terminate_process_tree can reliably signal
                # the whole tree via killpg.
                start_new_session=True,
            )
        finally:
            # Popen duplicates the handle for the child; safe to close our copy.
            log_file.close()

        with _get_project_lock(project):
            ann = load_annotations(pd)
            ann["pipeline_pid"] = proc.pid
            save_annotations(pd, ann)

    return {"status": PipelineStatus.Running, "project": project}
